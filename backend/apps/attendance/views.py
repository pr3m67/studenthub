from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
import math

import openpyxl
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from reportlab.pdfgen import canvas
from django.http import FileResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.models import User
from apps.accounts.permissions import IsTeacherOrAdmin
from .models import AttendanceRecord
from .serializers import AttendanceSerializer, TeacherAttendanceBoardSerializer
from .tasks import send_attendance_warning_email


def attendance_summary_for_user(user_id: str):
    grouped = defaultdict(lambda: {"present": 0, "absent": 0, "cancelled": 0, "total": 0})
    for record in AttendanceRecord.objects(user_id=user_id):
        item = grouped[record.subject]
        item[record.status] += 1
        if record.status != "cancelled":
            item["total"] += 1
    summary = []
    for subject, stats in grouped.items():
        attended = stats["present"]
        total = max(stats["total"], 1)
        percent = round(attended / total * 100, 2)
        needed = max(math.ceil(0.75 * total) - attended, 0)
        summary.append(
            {
                "subject": subject,
                "percentage": percent,
                "present": stats["present"],
                "absent": stats["absent"],
                "cancelled": stats["cancelled"],
                "risk": "danger" if percent < 60 else "warning" if percent < 75 else "safe",
                "needed_to_75": needed,
            }
        )
    return sorted(summary, key=lambda item: item["percentage"])


class AttendanceMarkView(APIView):
    def post(self, request):
        serializer = AttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data
        record = AttendanceRecord(
            user_id=payload["user_id"],
            subject=payload["subject"],
            date=payload["date"].strftime("%Y-%m-%d"),
            status=payload["status"],
            marked_by=payload.get("marked_by") or request.user.id,
        ).save()
        if payload["status"] != "cancelled":
            for item in attendance_summary_for_user(payload["user_id"]):
                if item["subject"] == payload["subject"] and item["percentage"] < 75:
                    send_attendance_warning_email.delay(payload["user_id"], payload["subject"])
                    break
        async_to_sync(get_channel_layer().group_send)(
            f"attendance_{payload['user_id']}",
            {"type": "attendance_message", "payload": {"subject": payload["subject"], "date": record.date, "status": record.status}},
        )
        return Response({"success": True, "data": AttendanceSerializer(record).data}, status=status.HTTP_201_CREATED)


class AttendanceBulkUploadView(APIView):
    permission_classes = [IsTeacherOrAdmin]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"success": False, "error": {"code": "ATT_001", "message": "Excel file is required.", "field_errors": {"file": ["Upload an .xlsx file."]}}}, status=400)
        workbook = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        sheet = workbook.active
        headers = [sheet.cell(row=1, column=index).value for index in range(1, 5)]
        if headers != ["Date", "Roll No", "Subject", "Status"]:
            return Response({"success": False, "error": {"code": "ATT_002", "message": "Invalid attendance template.", "field_errors": {"columns": ["Expected Date, Roll No, Subject, Status."]}}}, status=400)
        created = 0
        for row in range(2, sheet.max_row + 1):
            row_date = sheet.cell(row=row, column=1).value
            roll_no = str(sheet.cell(row=row, column=2).value or "").strip()
            subject = str(sheet.cell(row=row, column=3).value or "").strip()
            status_value = str(sheet.cell(row=row, column=4).value or "").strip().lower()
            if not roll_no or not subject or status_value not in {"present", "absent", "cancelled"}:
                continue
            user = User.objects(roll_no=roll_no).first()
            if not user:
                continue
            AttendanceRecord(user_id=user.id, subject=subject, date=row_date.strftime("%Y-%m-%d"), status=status_value, marked_by=request.user.id).save()
            created += 1
        return Response({"success": True, "count": created})


class AttendanceSummaryView(APIView):
    def get(self, request):
        summary = attendance_summary_for_user(request.user.id)
        return Response({"success": True, "data": summary})


class AttendanceCalendarView(APIView):
    def get(self, request):
        records = AttendanceRecord.objects(user_id=request.user.id).order_by("date")
        data = [{"date": record.date, "status": record.status, "subject": record.subject} for record in records]
        return Response({"success": True, "data": data})


class AttendancePredictionView(APIView):
    def get(self, request):
        results = []
        for item in attendance_summary_for_user(request.user.id):
            total = item["present"] + item["absent"]
            attended = item["present"]
            classes_needed = 0
            while total and (attended / total) < 0.75:
                total += 1
                attended += 1
                classes_needed += 1
            results.append({"subject": item["subject"], "classes_needed": classes_needed})
        return Response({"success": True, "data": results})


class AttendanceReportView(APIView):
    def get(self, request):
        subject = request.query_params.get("subject")
        summary = [item for item in attendance_summary_for_user(request.user.id) if not subject or item["subject"] == subject]
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        pdf.setTitle("StudentHub Attendance Report")
        y = 800
        pdf.drawString(50, y, f"Attendance Report for {request.user.name}")
        y -= 40
        for item in summary:
            pdf.drawString(50, y, f"{item['subject']}: {item['percentage']}% (P:{item['present']} A:{item['absent']} C:{item['cancelled']})")
            y -= 24
        pdf.save()
        buffer.seek(0)
        return FileResponse(buffer, filename="attendance_report.pdf", as_attachment=True)


class TeacherAttendanceBoardView(APIView):
    permission_classes = [IsTeacherOrAdmin]

    def get(self, request):
        division = request.query_params.get("division")
        batch = request.query_params.get("batch")
        students = User.objects(role="student")
        if division:
            students = students.filter(division=division)
        if batch:
            students = students.filter(batch=batch)
        data = [
            {
                "id": str(student.id),
                "name": student.name,
                "roll_no": student.roll_no,
                "division": student.division or "",
                "batch": student.batch or "",
                "selected_subjects": student.selected_subjects or [],
            }
            for student in students.order_by("roll_no")
        ]
        return Response({"success": True, "data": data})

    def post(self, request):
        serializer = TeacherAttendanceBoardSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data
        written = []
        for item in payload["records"]:
            student = User.objects(id=item.get("user_id"), role="student").first()
            if not student:
                continue
            record = AttendanceRecord(
                user_id=str(student.id),
                subject=payload["subject"],
                date=payload["date"].strftime("%Y-%m-%d"),
                status=item.get("status", "absent"),
                marked_by=str(request.user.id),
            ).save()
            written.append(str(record.id))
            async_to_sync(get_channel_layer().group_send)(
                f"attendance_{student.id}",
                {"type": "attendance_message", "payload": {"subject": payload["subject"], "date": record.date, "status": record.status}},
            )
        return Response({"success": True, "count": len(written), "data": written})
