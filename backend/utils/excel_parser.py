from __future__ import annotations

from io import BytesIO
from itertools import cycle
from typing import Any

import openpyxl
import pandas as pd
from rest_framework.exceptions import ValidationError

from utils.validators import validate_time_string

SUBJECT_COLORS = [
    "#2D3A3A",
    "#7C9A92",
    "#E8DCC8",
    "#7B8CDE",
    "#D98373",
    "#4CAF82",
    "#E8A838",
]

EXPECTED_COLUMNS = ["Day", "Start Time", "End Time", "Subject", "Teacher", "Venue"]
DAY_NAMES = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}
STOP_ROW_LABELS = {
    "faculty abbr.",
    "faculty name",
    "subject abbr.",
    "subject name",
    "pandit deendayal energy university",
    "school of technology",
    "b.tech - computer engineering",
}


def _normalize(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _color_map(subjects: list[str]) -> dict[str, str]:
    colors = cycle(SUBJECT_COLORS)
    mapping = {}
    for subject in subjects:
        if subject not in mapping:
            mapping[subject] = next(colors)
    return mapping


def _extract_lookup_pairs(sheet, label_column: int, label_value: str, key_column: int, value_column: int):
    start_row = None
    for row_index in range(1, sheet.max_row + 1):
        if _normalize(sheet.cell(row=row_index, column=label_column).value).strip().lower() == label_value.lower():
            start_row = row_index + 1
            break
    mapping = {}
    if not start_row:
        return mapping
    row = start_row
    while row <= sheet.max_row:
        key = _normalize(sheet.cell(row=row, column=key_column).value)
        value = _normalize(sheet.cell(row=row, column=value_column).value)
        if not key and not value:
            break
        if key and value:
            mapping[key] = value
        row += 1
    return mapping


def parse_simple_table(file_bytes: bytes) -> list[dict]:
    df = pd.read_excel(BytesIO(file_bytes), sheet_name="Timetable")
    missing = [column for column in EXPECTED_COLUMNS if column not in df.columns]
    if missing:
        raise ValidationError({"columns": [f"Missing columns: {', '.join(missing)}"]})
    entries = []
    color_mapping = _color_map(df["Subject"].dropna().astype(str).tolist())
    for index, row in df.iterrows():
        entry = {column: _normalize(row.get(column)) for column in EXPECTED_COLUMNS}
        if not all(entry.values()):
            raise ValidationError({f"row_{index + 2}": ["No empty values are allowed in timetable rows."]})
        validate_time_string(entry["Start Time"])
        validate_time_string(entry["End Time"])
        entries.append(
            {
                "day": entry["Day"],
                "start_time": entry["Start Time"],
                "end_time": entry["End Time"],
                "subject": entry["Subject"],
                "teacher": entry["Teacher"],
                "venue": entry["Venue"],
                "color_tag": color_mapping[entry["Subject"]],
                "slot_key": f"{entry['Day']}-{entry['Start Time']}-{entry['End Time']}",
                "batch_group": "",
                "raw_text": "",
                "source_type": "uploaded",
            }
        )
    return entries


def _parse_matrix_cell(cell_text: str, subject_lookup: dict[str, str], faculty_lookup: dict[str, str]):
    cleaned = _normalize(cell_text)
    if not cleaned:
        return None
    left, right = cleaned.rsplit(",", 1) if "," in cleaned else (cleaned, "")
    tokens = left.split()
    code = ""
    if "(" in left and ")" in left:
        code = left[left.find("(") + 1 : left.find(")")]
    batch_tokens = [token for token in tokens if token.upper().startswith("G")]
    teacher_abbr = right.split("-")[0].strip()
    teacher = faculty_lookup.get(teacher_abbr, teacher_abbr or "Faculty")
    venue = right.strip() or "Campus"
    subject = subject_lookup.get(code, code or left.strip())
    return {
        "subject": subject,
        "teacher": teacher,
        "venue": venue,
        "batch_group": ",".join(batch_tokens),
        "raw_text": cleaned,
    }


def parse_matrix_timetable(file_bytes: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    subject_lookup = _extract_lookup_pairs(sheet, label_column=5, label_value="Subject Abbr.", key_column=5, value_column=7)
    faculty_lookup = _extract_lookup_pairs(sheet, label_column=1, label_value="Faculty Abbr.", key_column=1, value_column=3)
    header_row = None
    for row_index in range(1, min(sheet.max_row, 12) + 1):
        if _normalize(sheet.cell(row=row_index, column=1).value).lower() == "day":
            header_row = row_index
            break
    if not header_row:
        raise ValidationError({"sheet": ["Could not find timetable header row."]})

    slots = {}
    for column in range(2, sheet.max_column + 1):
        value = _normalize(sheet.cell(row=header_row, column=column).value)
        if "-" in value:
            start, end = [part.strip() for part in value.split("-", 1)]
            validate_time_string(start)
            validate_time_string(end)
            slots[column] = (start, end)

    raw_entries = []
    current_day = ""
    for row in range(header_row + 1, sheet.max_row + 1):
        day_value = _normalize(sheet.cell(row=row, column=1).value)
        normalized_day = day_value.lower()
        if normalized_day in STOP_ROW_LABELS:
            break
        if day_value:
            if normalized_day not in DAY_NAMES:
                continue
            current_day = day_value
        if not current_day:
            continue
        for column, (start_time, end_time) in slots.items():
            parsed = _parse_matrix_cell(sheet.cell(row=row, column=column).value, subject_lookup, faculty_lookup)
            if parsed:
                raw_entries.append(
                    {
                        "day": current_day,
                        "start_time": start_time,
                        "end_time": end_time,
                        "slot_key": f"{current_day}-{start_time}-{end_time}",
                        "source_type": "uploaded",
                        **parsed,
                    }
                )
    color_mapping = _color_map([entry["subject"] for entry in raw_entries])
    for entry in raw_entries:
        entry["color_tag"] = color_mapping[entry["subject"]]
    return raw_entries


def extract_subject_catalog(file_bytes: bytes) -> list[str]:
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name="Timetable")
        if "Subject" in df.columns:
            return sorted({str(item).strip() for item in df["Subject"].dropna().tolist() if str(item).strip()})
    except Exception:
        pass

    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    subject_lookup = _extract_lookup_pairs(sheet, label_column=5, label_value="Subject Abbr.", key_column=5, value_column=7)
    excluded = {"CD CELL", "CDC"}
    return sorted({value.strip() for value in subject_lookup.values() if value and value.strip() and value.strip().upper() not in excluded})


def parse_timetable_file(file_bytes: bytes) -> list[dict]:
    try:
        return parse_simple_table(file_bytes)
    except Exception:
        return parse_matrix_timetable(file_bytes)
